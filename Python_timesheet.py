#Load Methon or libraly to program
from openpyxl import *
from tkinter import *
import tkinter.font as font
from tkinter import messagebox
from tkinter import colorchooser
from tkinter import filedialog
from tkinter import StringVar
from openpyxl.styles import PatternFill
from tkinter import ttk
import tkinter as tk
import pywin32_system32 as pywin
from getpass import *
import os
import time
from openpyxl.styles import PatternFill
import datetime
from tkcalendar import DateEntry

# ตัวแปรหลัก ต่างๆ
excel = Tk()

tab_1=ttk.Notebook(excel)
tab_1.pack(fill='both',expand='yes')
#load workbook----------------------------------------------------------------------------------------------------------
wd = load_workbook('Ratchanon_TimeSheet.xlsx')
wa = wd.active

wd_2 = load_workbook('Expense Form_Ratchanon.xlsx')
wa_2 = wd_2.active

wd_3 = load_workbook('Ratchanon_TimeSheet.xlsx')
wa_3 = wd_3.active

style_1 = ttk.Style()
style_1.configure("Treeview",
                  background = "silver",
                  foreground="back",
                  rowheight=25)

style_1.map("Treeview",background=[('selected','green')])

#tkinter
Timesheet_1 = Frame(tab_1)
Timesheet_1.pack(fill='both', expand='yes')
Expense_1 = Frame(tab_1)
Expense_1.pack()
#f1.grid(row=0, column=0)

x1=IntVar()
#btn_test = Button(excel, text='Submit', command=lambda :x1 ).pack()
#Label Frame 1 for timesheet
w1 = LabelFrame(Timesheet_1, text='DATA SHOW')
w2 = LabelFrame(Timesheet_1, text='CONTROL DATA')

#Label Frame 2 for expense
w3 = LabelFrame(Expense_1, text="DATA SHOW")
w4 = LabelFrame(Expense_1, text="CONTROL DATA")

#add tabb
tab_1.add(Timesheet_1, text="Time Sheet")
tab_1.add(Expense_1, text="Expense")
#frame in screen
w1.pack(fill='both', expand='yes', padx=10, pady=10)
w2.pack(fill='both', expand='yes', padx=10, pady=10)

w3.pack(fill='both', expand='yes',padx=10,pady=10)
w4.pack(fill='both', expand='yes',padx=10,pady=10)

w2.option_add("*Font","impact 15")
#    show_view=ttk.Treeview(w1, columns=(1,2,3,4), show='headings', height=20)
#    show_view.heading(1, text='Day(Sun-Sat)',anchor='c')
#    show_view.heading(2, text='Day')
#    show_view.heading(3, text='Project Name')
#    show_view.heading(4, text='Remark')
#    show_view.pack(padx=5, pady=5)

#treeview W1 frame and Treeview Timesheet
treeview_1=ttk.Treeview(w1)

treeview_1["columns"]=("Excel","NO","Date","Day","PN","Remark")
treeview_1.column("#0",width=0,minwidth=150,stretch=tk.NO)
treeview_1.column("Excel",width=0,minwidth=150,stretch=tk.NO)
treeview_1.column("NO",width=30,minwidth=150, stretch=tk.NO)
treeview_1.column("Date",width=50, minwidth=150, stretch=tk.NO)
treeview_1.column("Day", width=50,minwidth=150, stretch=tk.NO)
treeview_1.column("PN", width=270,minwidth=150, stretch=tk.NO)
treeview_1.column("Remark", width=270,minwidth=150, stretch=tk.NO)

treeview_1.heading("#0",text="")
treeview_1.heading("Excel", text="Ex Number",anchor=tk.W)
treeview_1.heading("NO", text="NO",anchor=tk.W)
treeview_1.heading("Date", text="Date",anchor=tk.W)
treeview_1.heading("Day",text="Day", anchor=tk.W)
treeview_1.heading("PN", text="Project Name",anchor=tk.W)
treeview_1.heading("Remark", text="Remark",anchor=tk.W)
treeview_1.pack()

#treeview frame2 and Treeview Expense
treeview_2=ttk.Treeview(w3)

treeview_2["columns"]=("Excel","NO","Day","Month","Description","Type","Local_Currenct" )
treeview_2.column("#0",width=0,minwidth=150,stretch=tk.NO) #value[index]
treeview_2.column("Excel",width=0,minwidth=150,stretch=tk.NO) #value[0]
treeview_2.column("NO",width=30,minwidth=150, stretch=tk.NO) #value[1]
treeview_2.column("Day",width=100, minwidth=150, stretch=tk.NO) #vale[2]
treeview_2.column("Month", width=100,minwidth=150, stretch=tk.NO) #value[3]
treeview_2.column("Description", width=200,minwidth=150, stretch=tk.NO) #value[4]
treeview_2.column("Type", width=100,minwidth=50, stretch=tk.NO) #value[5]
treeview_2.column("Local_Currenct", width=140, stretch=tk.NO) #value[6]

treeview_2.heading("#0",text="")
treeview_2.heading("Excel", text="Ex Number",anchor=tk.W)
treeview_2.heading("NO", text="NO",anchor=tk.CENTER)
treeview_2.heading("Day", text="Day",anchor=tk.CENTER)
treeview_2.heading("Month",text="Month", anchor=tk.CENTER)
treeview_2.heading("Description", text="Desciption",anchor=tk.CENTER)
treeview_2.heading("Type", text="Type",anchor=tk.CENTER)
treeview_2.heading("Local_Currenct",text="Local Currenct", anchor=tk.CENTER)
treeview_2.pack(fill="both",padx=75,pady=15)

treeview_1.tag_configure('Sun-Sat', background="red")
treeview_1.tag_configure('Normal',background="white")

#data to test insert to treeview
data_1= [
      [1,"1","FRI","HSBC Project","Test ESS"],
      [2,"2", "MON", "HSBC Project", "Test ESS"],
      [3,"3", "THU", "HSBC Project", "Test ESS"],
      [4,"4", "WED", "HSBC Project", "Test ESS"]
]

l_Date=Label(w2, text="DATE INPUT :",width=20).grid(row=0,column=0)
l_day=Label(w2,text="Day Input    :",width=20).grid(row=1,column=0)
l_ProjectN=Label(w2,text="Project Name :",width=20).grid(row=2,column=0)
l_Remark=Label(w2,text="Remark Input :",width=20).grid(row=3,column=0)

#text_Date=Entry(w2)
#text_Date.grid(row=0,column=1)
#text_Day= Entry(w2)
#text_Day.grid(row=1, column=1)
#text_ProjectN= Entry(w2)
#text_ProjectN.grid(row=2, column=1)
#text_Remark= Entry(w2)
#text_Remark.grid(row=3, column=1)

#option menu Mon
Select_Day=["",'Mon','Tue','Wed','Thu','Fri','Sat','Sun']
select_1= StringVar()
select_1.set('Mon')
Select_Day_1=OptionMenu(w2,select_1,*Select_Day)
Select_Day_1.grid(row=1,column=1,sticky='nsew')

#opton menu Project
Select_Bank=["",'TBANK PROJECT','HSBC PROJECT','KBANK PROJECT','TMB PROJECT']
select_2= StringVar()
select_2.set('HSBC PROJECT')
Select_Bank_1=OptionMenu(w2,select_2,*Select_Bank)
Select_Bank_1.grid(row=2,column=1,sticky='nsew')

#option menu Remark
Select_Remark=['Test UAT ESS','Export Data','Leave','Sick Leave','Holiday','']
select_3= StringVar()
select_3.set('TEST UAT ESS')
Select_Remark_1=OptionMenu(w2,select_3,*Select_Remark)
Select_Remark_1.grid(row=3,column=1,sticky='nsew')

#option Date
Select_Date=["",1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31]
select_4= StringVar()
select_4.set(1)
Select_Date_1=OptionMenu(w2,select_4,*Select_Date)
Select_Date_1.grid(row=0,column=1,sticky='nsew')

#All global is the best!
global count_1
global count_NO
global count_row
global count_add_row
global count_excel_No
global count_color_leave
global count_day_1
global count_mount_1
count_mount_1 = int(wa.cell(row=4,column=17).value)
#print('Print:count_mount_1 :'+str(count_mount_1))
count_color_leave=10
count_excel_No=10
count_add_row=10
count_row=10
count_NO=1
count_1=1
count_day_1 =1
lb_1=Label(w2)
lb_1.grid(row=5,column=0)

lb_2=Label(w2, text="Mont: "+str(wa.cell(row=4, column=17).value))
lb_2.grid(row=6,column=0)
lb_Signatue_date=Label(w2, text=wa['B45'].value)
lb_Signatue_date.grid(row=5, column=1)

lb_3=Label(excel, text="Time :")
lb_3.pack(side=LEFT)
text_Signature_up=Entry(w2, text='')
text_Signature_up.grid(row=6, column=1)

#fynctuion update color for holiy or Leave and Sat,Sun
def leave_color():
    global count_color_leave

    for fill_color in range(10,41):
        check_value_1=wa.cell(row=count_color_leave, column=12).value
        check_value_2=wa.cell(row=count_color_leave, column=1).value

        if check_value_2 == "Sat":
            grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

            wa['A' + str(count_color_leave)].fill = grayFill
            wa['B' + str(count_color_leave)].fill = grayFill
            wa['C' + str(count_color_leave)].fill = grayFill
            wa['D' + str(count_color_leave)].fill = grayFill
            wa['E' + str(count_color_leave)].fill = grayFill
            wa['F' + str(count_color_leave)].fill = grayFill
            wa['G' + str(count_color_leave)].fill = grayFill
            wa['H' + str(count_color_leave)].fill = grayFill
            wa['I' + str(count_color_leave)].fill = grayFill
            wa['J' + str(count_color_leave)].fill = grayFill
            wa['K' + str(count_color_leave)].fill = grayFill
            wa['L' + str(count_color_leave)].fill = grayFill
            #print("IF 1: "+str(count_color_leave)+" : " + str(wa.cell(row=count_color_leave, column=12).value))
            count_color_leave+=1

        elif check_value_2 == "Sun":
            grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

            wa['A' + str(count_color_leave)].fill = grayFill
            wa['B' + str(count_color_leave)].fill = grayFill
            wa['C' + str(count_color_leave)].fill = grayFill
            wa['D' + str(count_color_leave)].fill = grayFill
            wa['E' + str(count_color_leave)].fill = grayFill
            wa['F' + str(count_color_leave)].fill = grayFill
            wa['G' + str(count_color_leave)].fill = grayFill
            wa['H' + str(count_color_leave)].fill = grayFill
            wa['I' + str(count_color_leave)].fill = grayFill
            wa['J' + str(count_color_leave)].fill = grayFill
            wa['K' + str(count_color_leave)].fill = grayFill
            wa['L' + str(count_color_leave)].fill = grayFill
            #print("IF 2: " +str(count_color_leave)+" : " + str(wa.cell(row=count_color_leave, column=12).value))
            count_color_leave+=1

        elif check_value_1 == "Holiday" or check_value_1 == "Leave" or check_value_1 =="Sick Leave":
            grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

            wa['A' + str(count_color_leave)].fill = grayFill
            wa['B' + str(count_color_leave)].fill = grayFill
            wa['C' + str(count_color_leave)].fill = grayFill
            wa['D' + str(count_color_leave)].fill = grayFill
            wa['E' + str(count_color_leave)].fill = grayFill
            wa['F' + str(count_color_leave)].fill = grayFill
            wa['G' + str(count_color_leave)].fill = grayFill
            wa['H' + str(count_color_leave)].fill = grayFill
            wa['I' + str(count_color_leave)].fill = grayFill
            wa['J' + str(count_color_leave)].fill = grayFill
            wa['K' + str(count_color_leave)].fill = grayFill
            wa['L' + str(count_color_leave)].fill = grayFill
            #print("IF 3: "+str(count_color_leave)+" : " + str(wa.cell(row=count_color_leave, column=12).value))
            count_color_leave+=1

        elif check_value_1 == "Sick Leave":
            grayFill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

            wa['A' + str(count_color_leave)].fill = grayFill
            wa['B' + str(count_color_leave)].fill = grayFill
            wa['C' + str(count_color_leave)].fill = grayFill
            wa['D' + str(count_color_leave)].fill = grayFill
            wa['E' + str(count_color_leave)].fill = grayFill
            wa['F' + str(count_color_leave)].fill = grayFill
            wa['G' + str(count_color_leave)].fill = grayFill
            wa['H' + str(count_color_leave)].fill = grayFill
            wa['I' + str(count_color_leave)].fill = grayFill
            wa['J' + str(count_color_leave)].fill = grayFill
            wa['K' + str(count_color_leave)].fill = grayFill
            wa['L' + str(count_color_leave)].fill = grayFill
            #print("IF 4: "+str(count_color_leave)+" : " + str(wa.cell(row=count_color_leave, column=12).value))
            count_color_leave+=1

        elif count_color_leave >= 41:
            check_value_1=10

        elif wa.cell(row=count_color_leave, column=12).value != "Sick Leave"or"Holiday"or None:
            No_Fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

            wa['A' + str(count_color_leave)].fill = No_Fill
            wa['B' + str(count_color_leave)].fill = No_Fill
            wa['C' + str(count_color_leave)].fill = No_Fill
            wa['D' + str(count_color_leave)].fill = No_Fill
            wa['E' + str(count_color_leave)].fill = No_Fill
            wa['F' + str(count_color_leave)].fill = No_Fill
            wa['G' + str(count_color_leave)].fill = No_Fill
            wa['H' + str(count_color_leave)].fill = No_Fill
            wa['I' + str(count_color_leave)].fill = No_Fill
            wa['J' + str(count_color_leave)].fill = No_Fill
            wa['K' + str(count_color_leave)].fill = No_Fill
            wa['L' + str(count_color_leave)].fill = No_Fill
            #print("else :" +str(count_color_leave)+" : " + str(wa.cell(row=count_color_leave, column=12).value))
            count_color_leave+=1

#date Time function
def clock():
    h_1=time.strftime("%H")
    m_1=time.strftime("%M")
    s_1=time.strftime("%S")

    day_07=time.strftime("%a")
    mount_1=time.strftime("%B")
    year_1 = time.strftime("%Y")
    lb_3.config(text="Time: " + h_1 + ":" + m_1 + ":" + s_1 + "  " + day_07 + " " +mount_1+" "+year_1)

    lb_3.after(1000, clock)

#finction add data to Treevire
def add_DATA():
    global count_1
    global count_NO
    global count_add_row
    global count_excel_No
    global count_day_1

    cal_3 = datetime.datetime(2021, wa.cell(row=4,column=17).value ,count_day_1)
    cal3_result=cal_3.strftime('%a')

    if count_NO <= 31 and int(wa.cell(row=4,column=17).value == 2):
        treeview_1.insert("", 'end', iid=count_1, values=(count_excel_No,count_NO,count_day_1,cal3_result, select_2.get(), select_3.get()))
        wa['A'+str(count_add_row)]=cal3_result
        wa['B'+str(count_add_row)]=count_day_1
        wa['K'+str(count_add_row)]=select_2.get()
        wa['L'+str(count_add_row)]=select_3.get()
        #print(wa['K'+str(count_add_row)].value)
        #print(wa['L'+str(count_add_row)].value)
        count_1 +=1
        count_NO +=1
        count_add_row+=1
        count_excel_No+=1
        if count_day_1 <= 27:
            count_day_1 += 1
        elif count_1 >=31:
            count_1-=1
            lb_1.config(text="Error Treeview Add to over line 30")

    elif count_NO <= 31 and int(wa.cell(row=4,column=17).value in (1,3,5,7,8,12)):
        treeview_1.insert("", 'end', iid=count_1, values=(count_excel_No,count_NO,count_day_1,cal3_result, select_2.get(), select_3.get()))
        wa['A'+str(count_add_row)]=cal3_result
        wa['B'+str(count_add_row)]=count_day_1
        wa['K'+str(count_add_row)]=select_2.get()
        wa['L'+str(count_add_row)]=select_3.get()
        #print(wa['K'+str(count_add_row)].value)
        #print(wa['L'+str(count_add_row)].value)
        count_1 +=1
        count_NO +=1
        count_add_row+=1
        count_excel_No+=1
        if count_day_1 <= 31:
            count_day_1 += 1
        elif count_1 >=31:
            count_1-=1
            lb_1.config(text="Error Treeview Add to over line 30")

    elif count_NO <= 31 and int(wa.cell(row=4,column=17).value in (4,6,9,11)):
        treeview_1.insert("", 'end', iid=count_1, values=(count_excel_No,count_NO,count_day_1,cal3_result, select_2.get(), select_3.get()))
        wa['A'+str(count_add_row)]=cal3_result
        wa['B'+str(count_add_row)]=count_day_1
        wa['K'+str(count_add_row)]=select_2.get()
        wa['L'+str(count_add_row)]=select_3.get()
        #print(wa['K'+str(count_add_row)].value)
        #print(wa['L'+str(count_add_row)].value)
        count_1 +=1
        count_NO +=1
        count_add_row+=1
        count_excel_No+=1
        if count_day_1 <= 29:
            count_day_1 += 1
        elif count_1 >=32:
            count_1-=1
            lb_1.config(text="Error Treeview Add to over line 30")
    #elif count_day_1 == ValueError:
     #   count_day_1 = select_4.get()

#function update data to Treeview and update to excel: Remark[select record and change optontion and clike button Update Data]
def edit_DATA():
    global count_add_row
    edit_1=treeview_1.focus()
    value_1=treeview_1.item(edit_1, 'values')
    cell_ex1=wa['K'+str(count_excel_No)]
    treeview_1.item(edit_1, values=(value_1[0],str(edit_1),value_1[2], value_1[3], select_2.get(), select_3.get()))
    #wa['A'+str(value_1[0])]=select_4.get()
    #wa['B'+str(value_1[0])]=select_3.get()
    wa['K'+str(value_1[0])]=select_2.get()
    wa['L'+str(value_1[0])]=select_3.get()
    #print(wa['K'+str(count_add_row)].value)
    #print(wa['L'+str(count_add_row)].value)
    #print(value_1[0])

#function add data from excel to Treeview add 1:1 data
def ADD_F_Ex():
    global count_excel_No
    global count_1
    global count_NO
    global count_row
    global count_day_1
    global count_mount_1
    cal_2 = datetime.datetime(2021, count_mount_1, count_day_1)

    for cell in range(1,2):
        if count_NO <= 31 and int(wa.cell(row=4,column=17).value == 2 ):
            cal_result_2 = cal_2.strftime('%a')
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,values=(count_excel_No,count_NO, cell_Date, cal_result_2, cell_Pro,cell_R))
            wa['K'+str(count_excel_No)]=cell_Pro
            wa['L'+str(count_excel_No)]=cell_R
            count_1+=1
            count_NO+=1
            count_row+=1
            if count_NO <= 31:
                count_excel_No+=1
                if count_day_1 <28:
                    count_day_1+=1

        elif count_NO <= 31 and int(wa.cell(row=4, column=17).value in (1,3,5,7,8,12)):
            cal_result_2 = cal_2.strftime('%a')
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,values=(count_excel_No, count_NO, cell_Date, cal_result_2, cell_Pro, cell_R))
            wa['K' + str(count_excel_No)] = cell_Pro
            wa['L' + str(count_excel_No)] = cell_R
            count_1 += 1
            count_NO += 1
            count_row += 1
            if count_NO <= 31:
                count_excel_No += 1
                if count_day_1 < 30:
                    count_day_1 += 1

        elif count_NO <= 31 and int(wa.cell(row=4, column=17).value in (4,6,9,11)):
            cal_result_2 = cal_2.strftime('%a')
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,values=(count_excel_No, count_NO, cell_Date, cal_result_2, cell_Pro, cell_R))
            wa['K' + str(count_excel_No)] = cell_Pro
            wa['L' + str(count_excel_No)] = cell_R
            count_1 += 1
            count_NO += 1
            count_row += 1
            if count_NO <= 31:
                count_excel_No += 1
                if count_day_1 < 29:
                    count_day_1 += 1

#function add data from excel to Treeview K,L 10 - K,L 40
def ADD_F_All_Ex():
    global count_1
    global count_NO
    global count_row
    global count_excel_No
    global count_day_1
    global count_mount_1

    for cell in range(10,41):
        cal_1 = datetime.datetime(2021, int(wa.cell(row=4, column=17).value), count_day_1)
        if count_NO <= 31 and int(wa.cell(row=4, column=17).value == 2):
            cal_result = cal_1.strftime("%a")
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,values=(count_excel_No,count_NO, cell_Date,cal_result, cell_Pro,cell_R))
            wa['A'+str(count_excel_No)]=cal_result
            wa['B'+str(count_excel_No)]=count_day_1
            wa['K'+str(count_excel_No)]=cell_Pro
            wa['L'+str(count_excel_No)]=cell_R
            count_1+=1
            count_NO+=1
            count_row+=1
            count_excel_No+=1
#            print("Count Excel: "+str(count_NO))
            if count_NO <= 28:
                count_day_1 +=1
            elif count_excel_No == 41:
                break

        elif count_NO <= 31 and int(wa.cell(row=4, column=17).value in (1,3,5,7,8,12) ):
            cal_result = cal_1.strftime("%a")
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,
                                  values=(count_excel_No, count_NO, cell_Date, cal_result, cell_Pro, cell_R))
            wa['A' + str(count_excel_No)] = cal_result
            wa['K' + str(count_excel_No)] = cell_Pro
            wa['L' + str(count_excel_No)] = cell_R
            count_1 += 1
            count_NO += 1
            count_row += 1
            count_excel_No += 1
            #print("Count Excel: " + str(count_NO))
            if count_NO <= 31 :
                count_day_1 += 1
            elif count_excel_No == 41:
                break

        elif count_NO <= 31 and int(wa.cell(row=4, column=17).value in (4,6,9,11) ):
            cal_result = cal_1.strftime("%a")
            cell_R = wa.cell(row=count_row, column=12).value
            cell_Pro = wa.cell(row=count_row, column=11).value
            cell_Date = wa.cell(row=count_row, column=20).value
            treeview_1.insert("", 'end', iid=count_1,values=(count_excel_No, count_NO, cell_Date, cal_result, cell_Pro, cell_R))
            wa['A' + str(count_excel_No)] = cal_result
            wa['K' + str(count_excel_No)] = cell_Pro
            wa['L' + str(count_excel_No)] = cell_R
            count_1 += 1
            count_NO += 1
            count_row += 1
            count_excel_No += 1
            #print("Count Excel: " + str(count_NO))
            if count_NO <= 30:
                count_day_1 += 1
            elif count_excel_No == 41:
                break
            #print('Count ROW: '+str(count_row))
            #print('Count Excel: '+str(count_excel_No))



lb_m_1=Label(w1, text="Mont Time sheet :").pack(padx=5,pady=5,side=LEFT)
s_m_1=Entry(w1,width=3)
s_m_1.pack(side=LEFT,padx=10,pady=10)

#funcion selection mont for option selection
def select_mont_1():
    global count_day_1
    global count_excel_No
    global count_mount_1
    wa['Q4'] = int(s_m_1.get())
    lb_2.config(text="Mont: "+str(wa.cell(row=4,column=17).value))
    s_m_1.delete(0,END)

#function update DATE signature B:45 (Format: dd/mm/yyyy)
def up_sig_date():
    wa['B45']=str(text_Signature_up.get())
    #print(wa['B45'].value)
    lb_Signatue_date.config(text=wa['B45'].value)

    text_Signature_up.delete(0,END)

#save File and check file is duplicate
def save_file():
    leave_color()
    wd.save('Ratchanon_TimeSheet1.xlsx')


#remove all Treeview and no delelte data in file excel (Ratchanon_Timesheet.xsls)
def remove_all():
    global count_add_row
    global count_1
    global count_NO
    global count_excel_No
    global count_row
    global count_day_1
    for re_1 in treeview_1.get_children():
        treeview_1.delete(re_1)
        #wa['K'+str(count_add_row)]=""
        #wa['L'+str(count_add_row)]=""
        #count_add_row+=1
        count_1=0
        count_NO=1
        count_excel_No=10
        count_row=10
        count_day_1=1
        count_add_row = 10
        lb_1.config(text="")
        if count_add_row == wa['L40']:
            break
#remove data select 1:1 step and delete defalut timesheet cell K,L
def remove_one_data():
    global count_add_row
    edit_2 = treeview_1.focus()
    value_2 = treeview_1.item(edit_2, 'values')
    cell_ex1 = wa['K' + str(count_excel_No)]
    treeview_1.item(edit_2, values=(value_2[0],value_2[1],"","", "", ""))
    wa['A' + str(value_2[0])] = ""
    wa['B' + str(value_2[0])] = ""
    wa['K' + str(value_2[0])] = ""
    wa['L' + str(value_2[0])] = ""
    # print(wa['K'+str(count_add_row)].value)
    # print(wa['L'+str(count_add_row)].value)
    # print(value_1[0])


    #int(value_2[1]) --1
    #treeview_1.insert("", 'end', iid=value_2[1],values=(value_2[0], value_2[1], value_2[2], value_2[3], value_2[3], value_2[4]))
    #print(value_2[1])
    #wa['K'+str(value_2[0])]=""
    #wa['L'+str(value_2[0])]=""

#def select_mont_1():
#    global count_NO
#    for E21 in range(1,2):
#        lb_2.config(text="Mont: "+str(count_NO))
#        count_NO+=1
#test_002=10
#for te_1 in range(10,40):
#    print(wa.cell(row=test_002, column=23).value)
#    test_002+=1

clock()
btn_Add = Button(w2, text="ADD DATA", command=add_DATA).grid(row=4, column=0)
btn_Edit = Button(w2, text="Update DATA", command=edit_DATA).grid(row=4, column=1)
btn_Add_excel = Button(w2, text="ADD FROM EXCEL 1R",command=ADD_F_Ex).grid(row=3,column=2)
btn_Add_ALL_excel = Button(w2, text="ADD FROM ALL EXCEL",command=ADD_F_All_Ex).grid(row=4,column=2, padx=20)
btn_test=Button(w1, text="Submit",command=select_mont_1).pack(side=LEFT)
btn_save = Button(w2, text="Save File",command=save_file).grid(row=0,column=2,padx=20)
btn_save = Button(w2, text="UPDATE DATE",command=up_sig_date).grid(row=6, column=2)
btn_remove_Data=Button(w1, text="Delete Date Duplication",command=remove_one_data).pack(side=RIGHT , padx=10,pady=10)
btn_remove_Data=Button(w2, text="Remove Data All",command=remove_all).grid(row=2,column=2)
#btn_delete_File=Button(w2, text='Delete File Excel',command=os.remove("E:\Temp Python\Ratchanon_TimeSheet1.xlsx")).grid(row=7,column=0)
#***********************************************************************************************************************************************************************************************************************************
#Expense Dev Codding easy
global count_day_2
global count_excel_2
global count_mount_2
global count_NO_2
global count_2
global count_total_2

count_total_2 =0
count_day_2 = 1
count_excel_2 = 8
count_mount_2 = 2
Total_Price = 0
count_NO_2 = 1
count_2 = 1

lb2_Day = Label(w4, text="Day Input : ")
lb2_Day.grid(row=0,column=0,padx=5,pady=5)
lb2_Moth = Label(w4, text="Month Input : ")
lb2_Moth.grid(row=1,column=0,padx=5,pady=5)
lb2_Dscript = Label(w4, text="Description : ").grid(row=2,column=0,padx=5,pady=5)
lb2_Type = Label(w4, text ="Type Input : " ).grid(row=3,column=0,padx=5,pady=5)
lb2_money = Label(w4, text="Local Currenct Input : ").grid(row=4, column=0, padx=5, pady=5)

spinbox_1 = Spinbox(w4, from_=1, to=31, font=("Helvetica", 15))
spinbox_1.grid(row=0,column=1)

spinbox_2 = Spinbox(w4, from_=1, to=12,font=("Helvetica", 15) )
spinbox_2.grid(row=1, column=1)

text_descipt_1 =Entry(w4, text='',font=("Helvetica", 15))
text_descipt_1.grid(row=2, column=1, stick='we')

defalut_1 = StringVar(excel)
defalut_1.set(int(4))
spinbox_3 = Spinbox(w4, from_=1, to=6,font=("Helvetica", 15),textvariable=defalut_1 )
spinbox_3.grid(row=3, column=1)

text_Currency = Entry(w4, font=("Helvetica", 15))
text_Currency.grid(row=4, column=1, stick='we')

lb_total_currency_2 = Label(w3, text='Total Currency: ')
lb_total_currency_2.pack(side=LEFT)


def add_row():
    global count_2
    global count_NO_2
    global count_excel_2
    global count_mount_2
    global count_day_2
    global count_mount_1
    global count_total_2

    for add_row12 in range(1,2):
        cal_4=datetime.datetime(2021, wa_3.cell(row=4,column=17).value, count_day_2 )
        cal4_result= cal_4.strftime('%d')
        cal41_result = cal_4.strftime('%w')

        if cal41_result == str(1) or cal41_result == str(2) or cal41_result == str(3) or cal41_result == str(4) or cal41_result == str(5) :
            treeview_2.insert('', 'end',iid=count_2,values=(count_excel_2, count_NO_2, cal4_result,wa_3.cell(row=4, column=17).value, text_descipt_1.get(),spinbox_3.get(),text_Currency.get()))
            count_total_2 = count_total_2+int(text_Currency.get())
            lb_total_currency_2.config(text=f'Total Currency: {count_total_2} THB')
            wa_2['A'+str(count_excel_2)] = cal41_result
            wa_2['B'+str(count_excel_2)] = count_mount_1
            wa_2['D'+str(count_excel_2)] = text_descipt_1.get()
            wa_2['I'+str(count_excel_2)] = spinbox_3.get()
            wa_2['K'+str(count_excel_2)] = text_Currency.get()
            count_2 += 1
            count_excel_2 += 1
            count_NO_2 += 1
            if count_excel_2 == 29:
                break

        count_day_2 += 1

def ex_excel_all():
    global count_2
    global count_NO_2
    global count_excel_2
    global count_mount_2
    global count_day_2
    global count_mount_1
    global count_total_2

    for x_04 in range(1,30):
        treeview_2.insert('', 'end', iid=count_2, values=(count_excel_2, count_NO_2, wa_2.cell(row=count_excel_2,column=1).value, wa_3.cell(row=4,column=17).value, wa_2.cell(row=count_excel_2,column=4).value, wa_2.cell(row=count_excel_2,column=9).value, wa_2.cell(row=count_excel_2,column=11).value))
        count_excel_2+=1
        count_2+=1
        count_NO_2 +=1
        #print(wa_2.cell(row=count_excel_2,column=11))
        if count_excel_2 == 29:
            break

def update_excell():
    global count_2
    global count_NO_2
    global count_excel_2
    global count_mount_2
    global count_day_2
    global count_mount_1
    global count_total_2

    edit2_1=treeview_2.focus()
    value2_1=treeview_2.item(edit2_1, 'values')
    treeview_2.item(edit2_1, values=(value2_1[0],str(edit2_1),value2_1[2], value2_1[3], text_descipt_1.get(),spinbox_3.get() ,text_Currency.get()))
    #wa['A'+str(value_1[0])]=select_4.get()
    #wa['B'+str(value_1[0])]=select_3.get()
    #text_descipt_1.config(text=value2_1[4])
    wa_2['D'+str(value2_1[0])]= text_descipt_1.get()
    wa_2['K'+str(value2_1[0])]= int(text_Currency.get())
    #print(wa['K'+str(count_add_row)].value)
    #print(wa['L'+str(count_add_row)].value)
    #print(value_1[0])

def show_detail():
    select_5 = treeview_2.focus()
    cal_7 = treeview_2.item(select_5, 'values')

    #spinbox_1.insert(0, cal_7[2])
    #spinbox_2.insert(0, cal_7[3])
    text_descipt_1.insert(0, cal_7[4])
    text_Currency.insert(0, cal_7[5])


def delete_one():
    edit2_2 = treeview_2.focus()
    value2_2 = treeview_2.item(edit2_2, 'values')
    #cell2_ex1 = wa['K' + str(count_excel_No)]
    treeview_2.item(edit2_2, values=(value2_2[0],value2_2[1],"","", "", "",""))
    wa['A' + str(value2_2[0])] = ""
    wa['B' + str(value2_2[0])] = ""
    wa['D' + str(value2_2[0])] = ""
    wa['I' + str(value2_2[0])] = ""
    wa['K' + str(value2_2[0])] = ""

def delete_all():
    global count_2
    global count_NO_2
    global count_excel_2
    global count_mount_2
    global count_day_2
    global count_mount_1
    global count_total_2

    for re2_1 in treeview_2.get_children():
        treeview_2.delete(re2_1)
        #wa['K'+str(count_add_row)]=""
        #wa['L'+str(count_add_row)]=""
        #count_add_row+=1
        count_2=1
        count_NO_2=1
        count_excel_2=8
        count_day_2=1
        count_total_2 = 0
        count_add_row = 10
        #lb_1.config(text="")
        #if count_add_row == wa['L40']:
         #   break

def save_file_2():
    wd_2.save('Expense Form_Ratchanon_2.xlsx')

btn2_add_row_2 = Button(w4, text='ADD Column',command=add_row).grid(row=5, column=0,stick='nw',padx=12,pady=12)
btn2_update = Button(w4,text="UPDATE SECECT ROW",command=update_excell).grid(row=5,column=1,stick='nw',padx=12,pady=12)
btn2_Add_excell_all = Button(w4,text="ADD EXCEL FILE",command=ex_excel_all).grid(row=5,column=2,padx=12,pady=12, stick='nw')
btn_delete_one = Button(w4, text="Delete ROW",command=delete_one).grid(row=3,column=3,padx=12,pady=12,stick='nw')
btn_delete_All = Button(w4, text="Delete All",command=delete_all).grid(row=5,column=3,padx=12,pady=12,stick='nw')
btn_Save_2 = Button(w4,text="SAVE FILE",command=save_file_2).grid(row=4,column=3,padx=12,pady=12,stick='nw')
select_data = Button(w3, text="SELECT DATA",command=show_detail).pack(side=RIGHT)
excel.title("TIME SHEET")
excel.geometry("850x850")
excel.mainloop()
